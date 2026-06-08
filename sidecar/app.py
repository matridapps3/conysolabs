"""Conyso Bench — Python sidecar.

Standalone version of the LSS stats engine. NO LLM, NO S3, NO database.
Files are uploaded to a local data directory; computations run via the
existing stats modules; results are returned inline.

Endpoints:
  /parse/*               — Excel / CSV / PDF / screenshot parsers
  /wrangle/*             — outlier detection, column normalization
  /stats/capability      — Cp/Cpk/Pp/Ppk + Box-Cox transform
  /stats/hypothesis      — 27 hypothesis tests
  /stats/control_chart   — 9 charts + Western Electric + Nelson rules
  /stats/regression      — OLS, GLM, stepwise, best-subsets, logistic,
                           Poisson, nonlinear
  /stats/msa             — Gauge R&R
  /stats/doe             — factorial fit
  /stats/doe-design      — full / fractional / CCD / Box-Behnken / mixture /
                           Plackett-Burman / definitive screening designs
  /stats/response-surface — quadratic RSM fit + predicted optimum
  /stats/pareto          — Pareto with vital-few
  /stats/dpmo            — DPMO ↔ sigma level
  /stats/sample_size     — t / proportion / Cpk / ANOVA / regression
  /stats/predictive-cpk  — what-if simulator
  /stats/distribution-id — fit + rank candidate distributions
  /stats/reliability     — Weibull / exponential / Arrhenius
  /stats/multivariate    — PCA / k-means / LDA / hierarchical / Hotelling
  /stats/time_series     — exp smoothing / ARIMA / auto-ARIMA / decompose /
                           ACF/PACF / cross-correlation
  /stats/posthoc         — Tukey / Fisher / Games-Howell / Dunnett
  /stats/tolerance       — normal + non-parametric tolerance intervals
  /stats/probability     — distribution calculator + Q-Q plot
  /stats/graph           — boxplot / histogram / scatter / matrix / time
                           series / individual-value / run chart / multi-vari
  /stats/attribute-capability — binomial / Poisson capability
  /stats/anom            — Analysis of Means
  /stats/sixpack         — Capability Sixpack one-page report
  /stats/acceptance-sampling — single-sampling plan design + OC curve
  /stats/random-data     — random data generators (11 distributions)
"""

from __future__ import annotations

import io
import json
import math
import os
import uuid
from pathlib import Path

from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel


def _json_safe(obj):
    """Recursively replace NaN / +Inf / -Inf with None so the payload is
    STRICT, valid JSON. Python's json emits bare `NaN`/`Infinity` tokens by
    default (allow_nan=True) — these are invalid JSON and the browser's
    JSON.parse() rejects them, silently breaking the frontend on any analysis
    that produces a non-finite number (zero-variance capability, degenerate
    contingency tables, singular covariance matrices, etc.). Sanitising once
    at the serialization boundary covers every endpoint."""
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            # Safety net: an endpoint that forgot to convert its chart bytes to
            # a storage key (the easy-to-miss `_put_bytes` step) would otherwise
            # crash JSON serialization with a UnicodeDecodeError. Persist the
            # bytes here and expose a *_storage_key instead — so a wiring slip
            # degrades to "chart still works" rather than a 500.
            if isinstance(v, (bytes, bytearray)) and isinstance(k, str) and k.endswith("_png"):
                out[k[:-4] + "_storage_key"] = _put_bytes("charts", k[:-4] + ".png", bytes(v), "image/png")
            else:
                out[k] = _json_safe(v)
        return out
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, (bytes, bytearray)):
        # Stray binary not under a *_png key — store and return its key.
        return _put_bytes("charts", "artifact.png", bytes(obj), "image/png")
    return obj


class SafeJSONResponse(JSONResponse):
    """Default response class for the app — NaN/Inf become null, and we set
    allow_nan=False so any value that slips through sanitisation raises here
    (a loud server-side error) instead of shipping invalid JSON to the client."""
    def render(self, content) -> bytes:
        return json.dumps(
            _json_safe(content),
            ensure_ascii=False, allow_nan=False, separators=(",", ":"),
        ).encode("utf-8")

from parsers import excel as excel_parser
from parsers import pdf as pdf_parser
from parsers import csv_parser
from parsers import screenshot as screenshot_parser

# Imported for its side effect — sets matplotlib rcParams to the Conyso
# Bench editorial palette. Must run before any stats module renders a chart.
from stats import _theme  # noqa: F401
from stats import capability as capability_stat
from stats import hypothesis as hypothesis_stat
from stats import control_chart as control_chart_stat
from stats import regression as regression_stat
from stats import msa as msa_stat
from stats import doe as doe_stat
from stats import pareto as pareto_stat
from stats import dpmo as dpmo_stat
from stats import sample_size as sample_size_stat
from stats import validation as validation_stat
from stats import recommend as recommend_stat
from stats import simulation as simulation_stat
from stats import survey as survey_stat
from stats import text_pareto as text_pareto_stat
from stats import variance_budget as variance_budget_stat
from stats import flow as flow_stat
from stats import predictive_cpk as predictive_cpk_stat
from stats import distribution_id as distribution_id_stat
from stats import reliability as reliability_stat
from stats import multivariate as multivariate_stat
from stats import time_series as time_series_stat
from stats import posthoc as posthoc_stat
from stats import tolerance as tolerance_stat
from stats import probability as probability_stat
from stats import graphs as graphs_stat
from stats import attribute_capability as attribute_capability_stat
from stats import anom as anom_stat
from stats import sixpack as sixpack_stat
from stats import acceptance_sampling as acceptance_sampling_stat
from stats import random_data as random_data_stat
# New Bench-only modules (no Minitab equivalent for some of these):
from stats import agreement as agreement_stat
from stats import bootstrap as bootstrap_stat
from stats import correlation as correlation_stat
from stats import gage_linearity as gage_linearity_stat
from stats import survival as survival_stat
from stats import mixed_effects as mixed_effects_stat
from stats import preflight as preflight_stat
from stats import narrative as narrative_stat
from stats import followups as followups_stat
from stats import bayesian as bayesian_stat
from stats import quality_helpers as qh_stat
from wrangle import outliers as outliers_mod
from wrangle import transform as transform_mod

import functools as _functools
import inspect as _inspect
from fastapi.routing import APIRoute


def _finalize_artifacts(result):
    """Convert any `*_png` raw-bytes value in a result dict into a stored
    artifact + `*_storage_key`. Centralises the easy-to-forget step so an
    endpoint that returns its compute() dict directly can never leak binary
    into JSON (which crashes FastAPI's encoder → opaque 500). Applies to every
    route via ArtifactRoute below."""
    if isinstance(result, dict):
        for k in list(result.keys()):
            v = result[k]
            if isinstance(k, str) and k.endswith("_png") and isinstance(v, (bytes, bytearray)):
                base = k[:-4]                       # "chart_png" → "chart"
                result.pop(k)
                result[base + "_storage_key"] = _put_bytes("charts", base + ".png", bytes(v), "image/png")
    return result


class ArtifactRoute(APIRoute):
    """Wraps every endpoint so its return value is finalized (chart bytes →
    storage keys) before FastAPI serializes it."""
    def __init__(self, path, endpoint, **kw):
        if _inspect.iscoroutinefunction(endpoint):
            @_functools.wraps(endpoint)
            async def wrapped(*a, **k):
                return _finalize_artifacts(await endpoint(*a, **k))
        else:
            @_functools.wraps(endpoint)
            def wrapped(*a, **k):
                return _finalize_artifacts(endpoint(*a, **k))
        super().__init__(path, wrapped, **kw)


app = FastAPI(title="conyso-bench-sidecar", default_response_class=SafeJSONResponse)
app.router.route_class = ArtifactRoute   # must be set before any @app.post runs


@app.exception_handler(ValueError)
async def _value_error_handler(request, exc):
    # Stats functions raise ValueError for bad-but-plausible user input (wrong
    # column, too few levels, non-positive times, …). Without this, every such
    # case surfaced as an opaque 500 "Internal Server Error"; now the user gets
    # a clean 400 with the actual reason.
    return SafeJSONResponse(status_code=400, content={"detail": str(exc)})


@app.exception_handler(KeyError)
async def _key_error_handler(request, exc):
    # A KeyError from pandas almost always means the user pointed an analysis at
    # a column that isn't in the dataset. Surface it as a 400 with a clear hint
    # instead of a 500. (exc.args[0] is the missing key.)
    missing = exc.args[0] if exc.args else str(exc)
    return SafeJSONResponse(status_code=400,
                            content={"detail": f"column or field not found: {missing}"})


# Matplotlib figure-leak guard. Chart functions create figures via pyplot's
# global state machine; if any raise between plt.subplots() and the _png()
# that closes them, the figure orphans in pyplot's registry and never frees.
# Under sustained errored traffic this OOMs the container. This middleware
# closes every open figure after each request — success OR failure — so a
# leak can never accumulate across requests.
@app.middleware("http")
async def _close_figures(request, call_next):
    try:
        return await call_next(request)
    finally:
        try:
            import matplotlib.pyplot as _plt
            _plt.close("all")
        except Exception:
            pass


# Local data directory replaces the S3 bucket in the full-product version.
DATA_DIR = Path(os.environ.get("DATA_DIR", "./data")).resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
(CHARTS_DIR := DATA_DIR / "charts").mkdir(exist_ok=True)
(ROWS_DIR := DATA_DIR / "rows").mkdir(exist_ok=True)
(FILES_DIR := DATA_DIR / "files").mkdir(exist_ok=True)


def _get_bytes(storage_key: str) -> bytes:
    p = (DATA_DIR / storage_key).resolve()
    if not str(p).startswith(str(DATA_DIR)):
        raise HTTPException(status_code=400, detail="invalid storage_key")
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"not found: {storage_key}")
    return p.read_bytes()


def _put_bytes(prefix: str, filename: str, data: bytes, content_type: str) -> str:
    sub = DATA_DIR / prefix / str(uuid.uuid4())
    sub.mkdir(parents=True, exist_ok=True)
    p = sub / filename
    p.write_bytes(data)
    return str(p.relative_to(DATA_DIR))


def _read_rows(storage_key: str):
    """Parquet/CSV rows file → pandas DataFrame."""
    import pandas as pd
    data = _get_bytes(storage_key)
    # Try parquet first (preferred by the materializer); fall back to CSV.
    try:
        return pd.read_parquet(io.BytesIO(data))
    except Exception:
        return pd.read_csv(io.BytesIO(data))


# ─── Parse endpoints ────────────────────────────────────────────────────

class ParseReq(BaseModel):
    storage_key: str


# Common parser-error → clean-400 wrapper. Without this, an empty CSV or
# corrupt Excel surfaced as a 500 with no detail and the user saw
# "sidecar /parse/csv 500:" — useless. The wrapper traps pandas parsing
# errors + our own ValueErrors and surfaces them as 400 with a concrete
# detail string.
def _parse_safely(parser_fn, stream, kind: str):
    import pandas.errors as _pe
    try:
        return parser_fn(stream)
    except (_pe.EmptyDataError, _pe.ParserError, ValueError) as e:
        raise HTTPException(status_code=400, detail=f"{kind}_parse_failed: {e}")


@app.post("/parse/excel")
def parse_excel(req: ParseReq):
    return _parse_safely(excel_parser.parse,
                          io.BytesIO(_get_bytes(req.storage_key)), "excel")


@app.post("/parse/csv")
def parse_csv(req: ParseReq):
    return _parse_safely(csv_parser.parse,
                          io.BytesIO(_get_bytes(req.storage_key)), "csv")


@app.post("/parse/pdf")
def parse_pdf(req: ParseReq):
    return _parse_safely(pdf_parser.parse,
                          io.BytesIO(_get_bytes(req.storage_key)), "pdf")


# ─── Dataset preview + row fetch (for Explore view + quality report) ───
#
# Once a dataset has been materialized to parquet, the SPA needs two things:
#   1. A small sample + quality diagnostics for the upload feedback card
#   2. Up to N rows for client-side visualization (histograms, scatter, etc.)
# Both share the same parquet read; expose them as separate endpoints so
# the cheap preview doesn't accidentally ship megabytes of rows.

class DatasetRowsReq(BaseModel):
    rows_storage_key: str
    limit: int | None = 5000          # cap so we don't OOM the browser
    columns: list[str] | None = None


@app.post("/dataset/rows")
def dataset_rows(req: DatasetRowsReq):
    import math
    df = _read_rows(req.rows_storage_key)
    if req.columns:
        cols = [c for c in req.columns if c in df.columns]
        df = df[cols] if cols else df
    n_total = int(len(df))
    if req.limit and n_total > req.limit:
        df = df.head(req.limit)
    # NaN → None for JSON.
    def clean(rec):
        return {k: (None if isinstance(v, float) and math.isnan(v) else v)
                for k, v in rec.items()}
    return {
        "rows": [clean(r) for r in df.to_dict(orient="records")],
        "n_total": n_total,
        "n_returned": int(len(df)),
        "truncated": bool(req.limit and n_total > req.limit),
    }


class DatasetPreviewReq(BaseModel):
    rows_storage_key: str
    n: int = 20


@app.post("/dataset/preview")
def dataset_preview(req: DatasetPreviewReq):
    """Cheap preview + per-column data-quality flags. The flags drive the
    upload-feedback card: "Column 'temperature' is 87% numeric but 13%
    text — consider cleaning entries like 'N/A' before running stats."
    """
    import math
    import re
    import pandas as pd
    df = _read_rows(req.rows_storage_key)
    n = int(len(df))
    sample = df.head(req.n)
    # NaN → None
    def _clean(rec):
        return {k: (None if isinstance(v, float) and math.isnan(v) else v)
                for k, v in rec.items()}
    preview_rows = [_clean(r) for r in sample.to_dict(orient="records")]

    columns = []
    for col in df.columns:
        s = df[col]
        n_null = int(s.isna().sum())
        n_unique = int(s.nunique(dropna=True))
        kind = str(s.dtype)
        is_numeric = ("int" in kind) or ("float" in kind)
        is_datetime = "datetime" in kind
        type_label = (
            "number" if is_numeric else
            "date"   if is_datetime else
            "boolean" if kind == "bool" else
            "text"
        )
        flags = []
        if n > 0:
            null_pct = n_null / n * 100
            uniq_pct = n_unique / n * 100
            # 1. Mostly-numeric text columns — usually mixed N/A / stray strings.
            if type_label == "text" and n_unique > 1:
                non_null = s.dropna().astype(str)
                if len(non_null):
                    coerced = pd.to_numeric(
                        non_null.str.replace(",", "", regex=False).str.strip(),
                        errors="coerce",
                    )
                    pct_numeric = float(coerced.notna().mean() * 100)
                    if 40 <= pct_numeric < 100:
                        flags.append({
                            "level": "warn",
                            "msg": f"{pct_numeric:.0f}% of values are numeric — the rest are text. Clean stray entries (e.g. 'N/A', '-', 'TBD') before running stats.",
                        })
                    # Look for date-like text columns
                    elif pct_numeric < 5:
                        date_like = sum(1 for v in non_null.head(50)
                                        if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", str(v)) or
                                           re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", str(v)))
                        if date_like >= 5:
                            flags.append({
                                "level": "info",
                                "msg": "Looks like dates stored as text — date-based analyses will treat this as a string column.",
                            })
            # 2. Constant column
            if n_unique <= 1:
                flags.append({"level": "warn",
                              "msg": "Only one unique value — this column carries no information."})
            # 3. High null rate
            if null_pct >= 30:
                flags.append({"level": "warn",
                              "msg": f"{null_pct:.0f}% missing — high-null columns can break some analyses."})
            elif null_pct >= 10:
                flags.append({"level": "info",
                              "msg": f"{null_pct:.0f}% missing — usable for most tests, but be aware."})
            # 4. ID-like columns
            if type_label == "text" and uniq_pct > 95 and n_unique > 20:
                flags.append({"level": "info",
                              "msg": "Looks like an ID column — probably not a useful grouping variable."})
            # 5. High-cardinality grouping
            if type_label == "text" and 20 < n_unique < n and uniq_pct > 50:
                flags.append({"level": "info",
                              "msg": f"{n_unique} distinct values — too many for ANOVA-style group comparisons."})
        columns.append({
            "name": col, "type": type_label, "dtype": kind,
            "n_null": n_null, "n_unique": n_unique,
            "min": (None if not is_numeric else (None if s.dropna().empty else float(s.min()))),
            "max": (None if not is_numeric else (None if s.dropna().empty else float(s.max()))),
            "mean": (None if not is_numeric else (None if s.dropna().empty else float(s.mean()))),
            "flags": flags,
        })

    overall_flags = []
    if all(c["type"] != "number" for c in columns):
        overall_flags.append({"level": "warn",
                              "msg": "No numeric columns detected — most LSS analyses need at least one."})
    if n < 5:
        overall_flags.append({"level": "warn",
                              "msg": f"Only {n} row(s) — most tests need at least 10–30 for meaningful results."})
    return {
        "n_total": n,
        "preview_rows": preview_rows,
        "columns": columns,
        "overall_flags": overall_flags,
    }


# ─── In-memory ingestion paths (no file upload step) ───────────────────

class ParseTextReq(BaseModel):
    text: str
    name: str | None = None


@app.post("/parse/text")
def parse_text(req: ParseTextReq):
    """Parse pasted tabular text (CSV / TSV from Excel copy / pipe-DSV) without
    requiring an upload step first. The smart csv_parser handles delimiter +
    encoding + skip-blank detection."""
    if not req.text or not req.text.strip():
        raise HTTPException(status_code=400, detail="text_parse_failed: text is empty")
    return _parse_safely(csv_parser.parse, io.StringIO(req.text), "text")


class ParseJsonReq(BaseModel):
    data: list[dict] | dict
    name: str | None = None


@app.post("/parse/json")
def parse_json(req: ParseJsonReq):
    """Accept a JSON array of records `[{col: val, ...}, ...]` or a
    column-oriented `{col: [v1, v2, ...]}` payload and turn it into the
    standard rows + schema shape."""
    import pandas as pd
    payload = req.data
    if isinstance(payload, dict):
        # Column-oriented: convert to records.
        if not payload:
            raise HTTPException(status_code=400, detail="json_parse_failed: empty object")
        try:
            df = pd.DataFrame(payload)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"json_parse_failed: {e}")
    elif isinstance(payload, list):
        if not payload:
            raise HTTPException(status_code=400, detail="json_parse_failed: empty list")
        if not all(isinstance(r, dict) for r in payload):
            raise HTTPException(status_code=400,
                                 detail="json_parse_failed: expected list of objects")
        df = pd.DataFrame(payload)
    else:
        raise HTTPException(status_code=400, detail="json_parse_failed: must be list or object")
    df = df.dropna(axis=1, how="all")
    if df.empty:
        raise HTTPException(status_code=400, detail="json_parse_failed: no rows after cleanup")
    # NaN → None
    def _clean(rec):
        import math
        return {k: (None if isinstance(v, float) and math.isnan(v) else v)
                for k, v in rec.items()}
    rows = [_clean(r) for r in df.to_dict(orient="records")]
    return {
        "kind": "json",
        "header": list(df.columns),
        "n_rows": int(len(df)),
        "rows": rows,
        "sample": rows[:20],
        "dtypes": {c: str(df[c].dtype) for c in df.columns},
    }


# ─── Wrangle ────────────────────────────────────────────────────────────

class OutliersReq(BaseModel):
    rows_storage_key: str
    column: str
    method: str = "iqr"   # 'iqr' | 'zscore'
    threshold: float | None = None


@app.post("/wrangle/outliers")
def wrangle_outliers(req: OutliersReq):
    return outliers_mod.detect(_read_rows(req.rows_storage_key),
                               column=req.column, method=req.method,
                               threshold=req.threshold)


# ─── Stats: core ────────────────────────────────────────────────────────

class StatsReq(BaseModel):
    rows_storage_key: str
    column: str | None = None
    group_col: str | None = None
    lsl: float | None = None
    usl: float | None = None
    target: float | None = None
    transform: str | None = None
    test: str | None = None
    kind: str | None = None       # for control charts
    subgroup_col: str | None = None
    phase_col: str | None = None
    # Test-specific
    mu0: float | None = None
    p0: float | None = None
    equal_var: bool | None = None
    column_b: str | None = None
    median0: float | None = None
    delta: float | None = None
    factor_a: str | None = None
    factor_b: str | None = None
    columns: list[str] | None = None
    center: str | None = None
    n_col: str | None = None
    n: int | None = None
    subgroup_size: int | None = None
    lam: float | None = None
    L: float | None = None
    k: float | None = None
    h: float | None = None
    w: int | None = None
    # rm_anova
    subject_col: str | None = None
    within: str | None = None
    # two-way ANOVA SS type
    ss_type: str | None = None


@app.post("/stats/capability")
def capability(req: StatsReq):
    result = capability_stat.compute(_read_rows(req.rows_storage_key),
                                     column=req.column, lsl=req.lsl, usl=req.usl,
                                     target=req.target, transform=req.transform,
                                     subgroup_col=req.subgroup_col)
    if "chart_png" in result:
        key = _put_bytes("charts", "capability.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


@app.post("/stats/hypothesis")
def hypothesis(req: StatsReq):
    # Pass through the relevant kwargs.
    kwargs = req.model_dump(exclude_none=True)
    kwargs.pop("rows_storage_key", None)
    return hypothesis_stat.compute(_read_rows(req.rows_storage_key),
                                   test=req.test, column=req.column,
                                   group_col=req.group_col,
                                   **{k: v for k, v in kwargs.items()
                                      if k not in ("test", "column", "group_col")})


@app.post("/stats/control_chart")
def control_chart(req: StatsReq):
    kwargs = req.model_dump(exclude_none=True)
    kwargs.pop("rows_storage_key", None)
    kwargs.pop("kind", None)            # passed explicitly below — avoid duplicate kwarg
    result = control_chart_stat.compute(_read_rows(req.rows_storage_key),
                                        kind=req.kind, **kwargs)
    if "chart_png" in result:
        key = _put_bytes("charts", "control.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class RegressionReq(BaseModel):
    rows_storage_key: str
    response: str
    predictors: list[str] | None = None
    method: str | None = None        # ols | glm | logistic | poisson | nonlinear | stepwise | best_subsets | ordinal_logit
    family: str | None = None        # glm: gaussian | binomial | poisson | gamma
    formula: str | None = None       # glm: optional statsmodels formula override
    predictor: str | None = None     # nonlinear: single predictor
    model: str | None = None         # nonlinear: exp_decay | logistic | power | asymptotic
    p0: list[float] | None = None    # nonlinear: initial guess
    direction: str | None = None     # stepwise: forward | backward | both
    alpha_in: float | None = None
    alpha_out: float | None = None
    max_terms: int | None = None     # best_subsets
    penalty: float | None = None     # ridge/lasso/elastic-net: α (None ⇒ CV-chosen)
    l1_ratio: float | None = None    # elastic-net mixing (0=ridge … 1=lasso)


@app.post("/stats/regression")
def regression(req: RegressionReq):
    df = _read_rows(req.rows_storage_key)
    method = req.method or "ols"
    if method == "ols":
        result = regression_stat.compute(df, response=req.response,
                                         predictors=req.predictors or [])
    elif method == "glm":
        formula = req.formula or f"{req.response} ~ " + " + ".join(req.predictors or [])
        result = regression_stat.glm(df, formula=formula, family=req.family or "gaussian")
    elif method == "logistic":
        result = regression_stat.logistic(df, response=req.response,
                                          predictors=req.predictors or [])
    elif method == "poisson":
        result = regression_stat.poisson_regression(df, response=req.response,
                                                    predictors=req.predictors or [])
    elif method == "nonlinear":
        if not req.predictor or not req.model:
            raise HTTPException(status_code=400, detail="nonlinear regression requires predictor + model")
        result = regression_stat.nonlinear_regression(df, response=req.response,
                                                      predictor=req.predictor,
                                                      model=req.model, p0=req.p0)
    elif method == "stepwise":
        result = regression_stat.stepwise(df, response=req.response,
                                          predictors=req.predictors or [],
                                          direction=req.direction or "both",
                                          alpha_in=req.alpha_in if req.alpha_in is not None else 0.05,
                                          alpha_out=req.alpha_out if req.alpha_out is not None else 0.10)
    elif method == "best_subsets":
        result = regression_stat.best_subsets(df, response=req.response,
                                              predictors=req.predictors or [],
                                              max_k=req.max_terms)
    elif method == "ordinal_logit":
        result = regression_stat.ordinal_logit(df, response=req.response,
                                               predictors=req.predictors or [])
    elif method == "robust":
        # `family` is overloaded as the M-estimator name when method == robust.
        result = regression_stat.robust(df, response=req.response,
                                        predictors=req.predictors or [],
                                        m_estimator=(req.family or "huber"))
    elif method == "quantile":
        # `alpha_in` is overloaded as the quantile q for quantile regression.
        # Default to median (q=0.5).
        q = req.alpha_in if (req.alpha_in is not None and 0 < req.alpha_in < 1) else 0.5
        result = regression_stat.quantile(df, response=req.response,
                                          predictors=req.predictors or [],
                                          q=q)
    elif method == "random_forest":
        # max_terms repurposed as max_depth; alpha_in as a fraction is too
        # narrow so just expose max_terms for tree depth.
        result = regression_stat.random_forest(df, response=req.response,
                                               predictors=req.predictors or [],
                                               max_depth=req.max_terms)
    elif method == "pls":
        result = regression_stat.pls(df, response=req.response,
                                     predictors=req.predictors or [],
                                     n_components=req.max_terms)
    elif method == "beta":
        result = regression_stat.beta_regression(df, response=req.response,
                                                 predictors=req.predictors or [])
    elif method == "spline":
        if not req.predictor:
            raise HTTPException(status_code=400, detail="spline regression needs predictor")
        result = regression_stat.spline_regression(df, response=req.response,
                                                   predictor=req.predictor,
                                                   n_knots=int(req.max_terms or 4))
    elif method in ("ridge", "lasso", "elastic_net"):
        result = regression_stat.regularized(df, response=req.response,
                                             predictors=req.predictors or [],
                                             method=method,
                                             alpha=req.penalty,
                                             l1_ratio=req.l1_ratio if req.l1_ratio is not None else 0.5)
    else:
        raise HTTPException(status_code=400, detail=f"unknown regression method: {method}")
    if "chart_png" in result:
        key = _put_bytes("charts", "regression.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class MsaReq(BaseModel):
    rows_storage_key: str
    measurement_col: str
    part_col: str
    operator_col: str
    tolerance: float | None = None
    design: str | None = None             # crossed (default) | nested | expanded
    factor_cols: list[str] | None = None  # expanded: additional variance sources


@app.post("/stats/msa")
def msa(req: MsaReq):
    df = _read_rows(req.rows_storage_key)
    design = req.design or "crossed"
    if design == "crossed":
        result = msa_stat.compute(df, measurement_col=req.measurement_col,
                                  part_col=req.part_col, operator_col=req.operator_col,
                                  tolerance=req.tolerance)
    elif design == "nested":
        result = msa_stat.compute_nested(df, measurement_col=req.measurement_col,
                                         part_col=req.part_col, operator_col=req.operator_col,
                                         tolerance=req.tolerance)
    elif design == "expanded":
        result = msa_stat.compute_expanded(df, measurement_col=req.measurement_col,
                                           part_col=req.part_col, operator_col=req.operator_col,
                                           factor_cols=req.factor_cols or [],
                                           tolerance=req.tolerance)
    else:
        raise HTTPException(status_code=400, detail=f"unknown GR&R design: {design}")
    if "chart_png" in result:
        key = _put_bytes("charts", "msa.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class DoeReq(BaseModel):
    rows_storage_key: str
    response: str
    factors: list[str]
    interactions: bool = True


@app.post("/stats/doe")
def doe(req: DoeReq):
    result = doe_stat.compute(_read_rows(req.rows_storage_key),
                              response=req.response, factors=req.factors,
                              interactions=req.interactions)
    if "chart_png" in result:
        key = _put_bytes("charts", "doe.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class ParetoReq(BaseModel):
    rows_storage_key: str
    category_col: str
    threshold_pct: float = 80


@app.post("/stats/pareto")
def pareto(req: ParetoReq):
    return pareto_stat.compute(_read_rows(req.rows_storage_key),
                               category_col=req.category_col,
                               threshold_pct=req.threshold_pct)


class DpmoReq(BaseModel):
    defects: int
    units: int
    opportunities_per_unit: int = 1
    apply_shift: bool = True


@app.post("/stats/dpmo")
def dpmo(req: DpmoReq):
    return dpmo_stat.compute(defects=req.defects, units=req.units,
                             opportunities_per_unit=req.opportunities_per_unit,
                             apply_shift=req.apply_shift)


class SampleSizeReq(BaseModel):
    kind: str
    delta: float | None = None
    sigma: float | None = None
    p1: float | None = None
    p2: float | None = None
    cpk_target: float | None = None
    cpk_estimate: float | None = None
    confidence: float | None = None
    alpha: float = 0.05
    power: float = 0.80
    two_sample: bool = False
    two_sided: bool = True
    k_groups: int | None = None
    effect_size_f: float | None = None
    effect_size_f2: float | None = None
    n_predictors: int | None = None
    # extended cases
    df_chi: int | None = None
    effect_size_w: float | None = None
    true_mean_diff: float | None = None
    hazard_ratio: float | None = None
    p_event: float | None = None
    allocation_ratio: float | None = None
    icc: float | None = None
    cluster_size: int | None = None
    n_required_infinite: int | None = None
    population_size: int | None = None
    sigma2_ratio: float | None = None
    r: float | None = None


@app.post("/stats/sample_size")
def sample_size(req: SampleSizeReq):
    # Required inputs per kind. Without these the calculator would compare a
    # missing value (None) against a number deep inside (e.g. `None <= 0`),
    # raising a TypeError that surfaces as a raw 500. Validate up front so an
    # incomplete form returns a clear 400 instead. (chi_square and fpc apply
    # their own defaults, and proportion_test handles a missing p2 cleanly.)
    REQUIRED = {
        "t_test":             ["delta", "sigma"],
        "proportion_test":    ["p1"],
        "cpk_validation":     ["cpk_target", "cpk_estimate"],
        "anova":              ["k_groups", "effect_size_f"],
        "regression":         ["n_predictors", "effect_size_f2"],
        "equivalence":        ["delta", "sigma"],
        "logrank":            ["hazard_ratio"],
        "cluster_randomized": ["delta", "sigma"],
        "variance_test":      ["sigma2_ratio"],
        "correlation":        ["r"],
    }
    missing = [f for f in REQUIRED.get(req.kind, []) if getattr(req, f) is None]
    if missing:
        raise HTTPException(status_code=400,
                            detail=f"{req.kind} sample size needs: {', '.join(missing)}.")
    if req.kind == "t_test":
        return sample_size_stat.t_test(delta=req.delta, sigma=req.sigma,
                                       alpha=req.alpha, power=req.power,
                                       two_sample=req.two_sample,
                                       two_sided=req.two_sided)
    if req.kind == "proportion_test":
        return sample_size_stat.proportion_test(p1=req.p1, p2=req.p2,
                                                alpha=req.alpha, power=req.power,
                                                two_sided=req.two_sided)
    if req.kind == "cpk_validation":
        return sample_size_stat.cpk_validation(cpk_target=req.cpk_target,
                                               cpk_estimate=req.cpk_estimate,
                                               confidence=req.confidence or 0.95)
    if req.kind == "anova":
        return sample_size_stat.anova(k_groups=req.k_groups,
                                      effect_size_f=req.effect_size_f,
                                      alpha=req.alpha, power=req.power)
    if req.kind == "regression":
        return sample_size_stat.regression(n_predictors=req.n_predictors,
                                           effect_size_f2=req.effect_size_f2,
                                           alpha=req.alpha, power=req.power)
    if req.kind == "chi_square":
        return sample_size_stat.chi_square(df_chi=req.df_chi or 1,
                                           effect_size_w=req.effect_size_w or 0.3,
                                           alpha=req.alpha, power=req.power)
    if req.kind == "equivalence":
        return sample_size_stat.equivalence_tost(
            delta=req.delta, sigma=req.sigma,
            true_mean_diff=req.true_mean_diff or 0.0,
            alpha=req.alpha, power=req.power, two_sample=req.two_sample)
    if req.kind == "logrank":
        return sample_size_stat.logrank(
            hazard_ratio=req.hazard_ratio,
            p_event=req.p_event if req.p_event is not None else 0.5,
            alpha=req.alpha, power=req.power, two_sided=req.two_sided,
            allocation_ratio=req.allocation_ratio or 1.0)
    if req.kind == "cluster_randomized":
        return sample_size_stat.cluster_randomized(
            delta=req.delta, sigma=req.sigma,
            icc=req.icc or 0.05,
            cluster_size=req.cluster_size or 10,
            alpha=req.alpha, power=req.power, two_sided=req.two_sided)
    if req.kind == "fpc":
        return sample_size_stat.finite_population_correction(
            n_required_infinite=req.n_required_infinite or 0,
            population_size=req.population_size or 0)
    if req.kind == "variance_test":
        return sample_size_stat.variance_test(
            sigma2_ratio=req.sigma2_ratio,
            alpha=req.alpha, power=req.power)
    if req.kind == "correlation":
        return sample_size_stat.correlation(
            r=req.r, alpha=req.alpha, power=req.power, two_sided=req.two_sided)
    raise HTTPException(status_code=400, detail=f"unknown kind: {req.kind}")


# ─── Stats: advanced ────────────────────────────────────────────────────

class PredictiveCpkReq(BaseModel):
    rows_storage_key: str
    column: str
    lsl: float | None = None
    usl: float | None = None
    scenarios: list[dict] | None = None


@app.post("/stats/predictive-cpk")
def predictive_cpk(req: PredictiveCpkReq):
    result = predictive_cpk_stat.compute(_read_rows(req.rows_storage_key),
                                         column=req.column,
                                         lsl=req.lsl, usl=req.usl,
                                         scenarios=req.scenarios)
    if "chart_png" in result:
        key = _put_bytes("charts", "predictive_cpk.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class DistributionIdReq(BaseModel):
    rows_storage_key: str
    column: str
    candidates: list[str] | None = None


@app.post("/stats/distribution-id")
def distribution_id(req: DistributionIdReq):
    return distribution_id_stat.compute(_read_rows(req.rows_storage_key),
                                        column=req.column,
                                        candidates=req.candidates)


class ReliabilityReq(BaseModel):
    rows_storage_key: str
    distribution: str = "weibull"
    time_col: str
    censor_col: str | None = None
    mission_times: list[float] | None = None
    temp_col_kelvin: str | None = None
    use_kelvin: float | None = None
    # Cox PH only:
    event_col: str | None = None
    predictors: list[str] | None = None


@app.post("/stats/reliability")
def reliability(req: ReliabilityReq):
    df = _read_rows(req.rows_storage_key)
    if req.distribution == "weibull":
        result = reliability_stat.weibull(df, time_col=req.time_col,
                                          censor_col=req.censor_col,
                                          mission_times=req.mission_times)
    elif req.distribution == "exponential":
        result = reliability_stat.exponential(df, time_col=req.time_col,
                                              censor_col=req.censor_col,
                                              mission_times=req.mission_times)
    elif req.distribution == "arrhenius":
        if not req.temp_col_kelvin:
            raise HTTPException(status_code=400, detail="arrhenius requires temp_col_kelvin")
        result = reliability_stat.arrhenius_acceleration(df,
                                                         time_col=req.time_col,
                                                         temp_col_kelvin=req.temp_col_kelvin,
                                                         censor_col=req.censor_col,
                                                         use_kelvin=req.use_kelvin)
    elif req.distribution == "lognormal":
        result = reliability_stat.lognormal(df, time_col=req.time_col,
                                            censor_col=req.censor_col,
                                            mission_times=req.mission_times)
    elif req.distribution == "gamma":
        result = reliability_stat.gamma(df, time_col=req.time_col,
                                        censor_col=req.censor_col,
                                        mission_times=req.mission_times)
    elif req.distribution == "log_logistic":
        result = reliability_stat.log_logistic(df, time_col=req.time_col,
                                               censor_col=req.censor_col,
                                               mission_times=req.mission_times)
    elif req.distribution == "smallest_extreme_value":
        result = reliability_stat.smallest_extreme_value(df, time_col=req.time_col,
                                                         censor_col=req.censor_col,
                                                         mission_times=req.mission_times)
    elif req.distribution == "largest_extreme_value":
        result = reliability_stat.largest_extreme_value(df, time_col=req.time_col,
                                                        censor_col=req.censor_col,
                                                        mission_times=req.mission_times)
    elif req.distribution == "gev":
        result = reliability_stat.gev(df, time_col=req.time_col,
                                      censor_col=req.censor_col,
                                      mission_times=req.mission_times)
    elif req.distribution == "cox_ph":
        if not req.event_col or not req.predictors:
            raise HTTPException(status_code=400, detail="cox_ph requires event_col + predictors")
        result = reliability_stat.cox_ph(df, time_col=req.time_col,
                                         event_col=req.event_col,
                                         predictors=req.predictors)
    elif req.distribution == "crow_amsaa":
        result = reliability_stat.crow_amsaa(df, time_col=req.time_col,
                                             failure_col=req.censor_col)
    elif req.distribution == "eyring":
        if not req.temp_col_kelvin:
            raise HTTPException(status_code=400, detail="eyring requires temp_col_kelvin as stress column")
        result = reliability_stat.eyring(df, time_col=req.time_col,
                                         stress_col=req.temp_col_kelvin,
                                         censor_col=req.censor_col,
                                         use_stress=req.use_kelvin)
    elif req.distribution == "inverse_power_law":
        if not req.temp_col_kelvin:
            raise HTTPException(status_code=400, detail="inverse_power_law requires stress column (use temp_col_kelvin as the stress field)")
        result = reliability_stat.inverse_power_law(df, time_col=req.time_col,
                                                     stress_col=req.temp_col_kelvin,
                                                     censor_col=req.censor_col,
                                                     use_stress=req.use_kelvin)
    else:
        raise HTTPException(status_code=400, detail=f"unknown distribution: {req.distribution}")
    if "chart_png" in result:
        key = _put_bytes("charts", "reliability.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class MultivariateReq(BaseModel):
    rows_storage_key: str
    method: str
    columns: list[str] | None = None
    k: int | None = None
    n_components: int | None = None
    standardize: bool = True
    class_col: str | None = None
    mu0: list[float] | None = None


@app.post("/stats/multivariate")
def multivariate(req: MultivariateReq):
    df = _read_rows(req.rows_storage_key)
    cols = req.columns or []
    if req.method == "pca":
        result = multivariate_stat.pca(df, columns=cols,
                                       n_components=req.n_components,
                                       standardize=req.standardize)
    elif req.method == "kmeans":
        result = multivariate_stat.kmeans(df, columns=cols, k=req.k,
                                          standardize=req.standardize)
    elif req.method == "lda":
        if not req.class_col:
            raise HTTPException(status_code=400, detail="lda requires class_col")
        result = multivariate_stat.lda(df, predictors=cols, class_col=req.class_col)
    elif req.method == "hierarchical":
        result = multivariate_stat.hierarchical_cluster(df, columns=cols,
                                                        n_clusters=req.k,
                                                        standardize=req.standardize)
    elif req.method == "hotelling":
        result = multivariate_stat.hotelling_t2(df, columns=cols, mu0=req.mu0)
    elif req.method == "manova":
        if not req.class_col:
            raise HTTPException(status_code=400, detail="manova requires class_col (the grouping factor)")
        result = multivariate_stat.manova(df, responses=cols, factor=req.class_col)
    elif req.method == "factor":
        result = multivariate_stat.factor_analysis(df, columns=cols,
                                                   n_factors=req.n_components,
                                                   rotation="varimax")
    else:
        raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")
    if "chart_png" in result:
        key = _put_bytes("charts", f"{req.method}.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class TimeSeriesReq(BaseModel):
    rows_storage_key: str
    method: str
    value_col: str
    time_col: str | None = None
    horizon: int = 12
    trend: str | None = None
    seasonal: str | None = None
    seasonal_periods: int | None = None
    p: int | None = None
    d: int | None = None
    q: int | None = None
    seasonal_order: list[int] | None = None
    period: int | None = None
    model: str | None = None
    nlags: int | None = None
    x_col: str | None = None
    y_col: str | None = None
    max_lag: int | None = None


@app.post("/stats/time_series")
def time_series(req: TimeSeriesReq):
    df = _read_rows(req.rows_storage_key)
    if req.method == "exp_smoothing":
        result = time_series_stat.exponential_smoothing(df, value_col=req.value_col,
                                                        time_col=req.time_col,
                                                        trend=req.trend, seasonal=req.seasonal,
                                                        seasonal_periods=req.seasonal_periods,
                                                        horizon=req.horizon)
    elif req.method == "arima":
        result = time_series_stat.arima(df, value_col=req.value_col, time_col=req.time_col,
                                        p=req.p or 1, d=req.d or 0, q=req.q or 0,
                                        seasonal_order=tuple(req.seasonal_order) if req.seasonal_order else None,
                                        horizon=req.horizon)
    elif req.method == "auto_arima":
        result = time_series_stat.auto_arima(df, value_col=req.value_col,
                                             time_col=req.time_col, horizon=req.horizon)
    elif req.method == "decompose":
        result = time_series_stat.decompose(df, value_col=req.value_col,
                                            time_col=req.time_col,
                                            period=req.period or 12,
                                            model=req.model or "additive")
    elif req.method == "acf_pacf":
        result = time_series_stat.acf_pacf(df, value_col=req.value_col,
                                           time_col=req.time_col, nlags=req.nlags or 20)
    elif req.method == "cross_correlation":
        if not req.x_col or not req.y_col:
            raise HTTPException(status_code=400, detail="cross_correlation requires x_col and y_col")
        result = time_series_stat.cross_correlation(df, x_col=req.x_col, y_col=req.y_col,
                                                    time_col=req.time_col,
                                                    max_lag=req.max_lag or 20)
    elif req.method == "changepoint":
        result = time_series_stat.changepoint(df, value_col=req.value_col,
                                              time_col=req.time_col)
    else:
        raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")
    if "chart_png" in result:
        key = _put_bytes("charts", f"{req.method}.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class DesignReq(BaseModel):
    design: str
    factors: list[str]
    generators: list[str] | None = None
    alpha: float | None = None
    center_runs: int | None = None
    # Taguchi-only:
    factors_with_levels: list[dict] | None = None   # [{name, levels: [...]}]
    array: str | None = None                         # 'L4'|'L8'|'L9'|'L12'|'L16'
    n_runs: int | None = None                        # optimal: run budget
    model: str | None = None                         # optimal: linear|interaction|quadratic
    criterion: str | None = None                     # optimal: D|I
    levels: int | None = None                        # optimal: candidate-grid levels


@app.post("/stats/doe-design")
def doe_design(req: DesignReq):
    if req.design == "full_factorial":
        return doe_stat.full_factorial_2k(req.factors)
    if req.design == "fractional_factorial":
        return doe_stat.fractional_factorial(req.factors, generators=req.generators)
    if req.design == "central_composite":
        return doe_stat.central_composite(req.factors, alpha=req.alpha,
                                          center_runs=req.center_runs if req.center_runs is not None else 4)
    if req.design == "box_behnken":
        return doe_stat.box_behnken(req.factors,
                                    center_runs=req.center_runs if req.center_runs is not None else 3)
    if req.design == "plackett_burman":
        return doe_stat.plackett_burman(req.factors)
    if req.design == "definitive_screening":
        return doe_stat.definitive_screening(req.factors)
    if req.design == "mixture_simplex_lattice":
        return doe_stat.mixture_simplex_lattice(req.factors, degree=req.center_runs or 2)
    if req.design == "mixture_simplex_centroid":
        return doe_stat.mixture_simplex_centroid(req.factors)
    if req.design == "taguchi":
        if not req.factors_with_levels:
            raise HTTPException(status_code=400,
                                detail="taguchi requires factors_with_levels: [{name, levels:[…]}]")
        return doe_stat.taguchi(req.factors_with_levels, array=req.array)
    if req.design in ("d_optimal", "i_optimal"):
        if not req.n_runs:
            raise HTTPException(status_code=400, detail="optimal design requires n_runs")
        return doe_stat.optimal_design(req.factors, n_runs=req.n_runs,
                                       model=req.model or "interaction",
                                       criterion="I" if req.design == "i_optimal" else "D",
                                       levels=req.levels or 3)
    raise HTTPException(status_code=400, detail=f"unknown design: {req.design}")


class DoePowerReq(BaseModel):
    n_runs: int
    n_factors: int
    effect_size: float
    alpha: float = 0.05
    model: str = "interaction"
    n_replicates: int = 1


@app.post("/stats/doe-power")
def doe_power(req: DoePowerReq):
    return doe_stat.factorial_power(n_runs=req.n_runs, n_factors=req.n_factors,
                                    effect_size=req.effect_size, alpha=req.alpha,
                                    model=req.model, n_replicates=req.n_replicates)


class PowerCurveReq(BaseModel):
    kind: str = "two_sample_t"        # one_sample_t | two_sample_t | two_proportions | anova
    effect_size: float | None = None
    delta: float | None = None
    sigma: float | None = None
    p1: float | None = None
    p2: float | None = None
    k_groups: int = 3
    alpha: float = 0.05
    power: float = 0.80
    two_sided: bool = True


@app.get("/validation/nist")
def validation_nist():
    return validation_stat.nist_strd()


class RecommendReq(BaseModel):
    phase: str
    dataset: dict | None = None
    history: list | None = None
    open_items: list | None = None


@app.post("/recommend")
def recommend(req: RecommendReq):
    return recommend_stat.recommend(phase=req.phase, dataset=req.dataset,
                                    history=req.history, open_items=req.open_items)


class MonteCarloReq(BaseModel):
    inputs: list
    transfer: dict
    n_runs: int = 10000
    lsl: float | None = None
    usl: float | None = None
    target: float | None = None
    seed: int = 20260531


@app.post("/stats/monte-carlo")
def monte_carlo(req: MonteCarloReq):
    return simulation_stat.monte_carlo(req.inputs, req.transfer, n_runs=req.n_runs,
                                       lsl=req.lsl, usl=req.usl, target=req.target, seed=req.seed)


class ToleranceStackReq(BaseModel):
    inputs: list
    method: str = "both"
    lsl: float | None = None
    usl: float | None = None


@app.post("/stats/tolerance-stack")
def tolerance_stack(req: ToleranceStackReq):
    return simulation_stat.tolerance_stack(req.inputs, method=req.method, lsl=req.lsl, usl=req.usl)


class SurveyReq(BaseModel):
    rows_storage_key: str
    items: list
    scale_min: int | None = None
    scale_max: int | None = None


@app.post("/stats/survey")
def survey(req: SurveyReq):
    return survey_stat.analyze(_read_rows(req.rows_storage_key), items=req.items,
                               scale_min=req.scale_min, scale_max=req.scale_max)


class TextParetoReq(BaseModel):
    rows_storage_key: str
    text_col: str
    top_n: int = 10
    themes: dict | None = None
    use_bigrams: bool = True
    threshold_pct: float = 80.0


@app.post("/stats/text-pareto")
def text_pareto(req: TextParetoReq):
    return text_pareto_stat.analyze(_read_rows(req.rows_storage_key), text_col=req.text_col,
                                    top_n=req.top_n, themes=req.themes,
                                    use_bigrams=req.use_bigrams, threshold_pct=req.threshold_pct)


class VarianceBudgetReq(BaseModel):
    rows_storage_key: str
    response: str
    factors: list


@app.post("/stats/variance-budget")
def variance_budget(req: VarianceBudgetReq):
    return variance_budget_stat.analyze(_read_rows(req.rows_storage_key),
                                        response=req.response, factors=req.factors)


class CycleTimeReq(BaseModel):
    rows_storage_key: str
    time_col: str | None = None
    start_col: str | None = None
    end_col: str | None = None


@app.post("/stats/cycle-time")
def cycle_time(req: CycleTimeReq):
    return flow_stat.cycle_time(_read_rows(req.rows_storage_key), time_col=req.time_col,
                                start_col=req.start_col, end_col=req.end_col)


class DeliveryForecastReq(BaseModel):
    rows_storage_key: str
    throughput_col: str
    backlog: int | None = None
    horizon: int | None = None


@app.post("/stats/delivery-forecast")
def delivery_forecast(req: DeliveryForecastReq):
    import pandas as pd
    df = _read_rows(req.rows_storage_key)
    if req.throughput_col not in df.columns:
        raise ValueError(f"column '{req.throughput_col}' not in dataset")
    tp = pd.to_numeric(df[req.throughput_col], errors="coerce").dropna().tolist()
    return flow_stat.delivery_forecast(tp, backlog=req.backlog, horizon=req.horizon)


class LittlesLawReq(BaseModel):
    wip: float | None = None
    throughput: float | None = None
    cycle_time: float | None = None


@app.post("/stats/littles-law")
def littles_law(req: LittlesLawReq):
    return flow_stat.littles_law(wip=req.wip, throughput=req.throughput, cycle_time=req.cycle_time)


@app.post("/stats/power-curve")
def power_curve(req: PowerCurveReq):
    result = sample_size_stat.power_curve(
        kind=req.kind, effect_size=req.effect_size, delta=req.delta,
        sigma=req.sigma, p1=req.p1, p2=req.p2, k_groups=req.k_groups,
        alpha=req.alpha, power=req.power, two_sided=req.two_sided)
    if "chart_png" in result:
        key = _put_bytes("charts", "power_curve.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class DoeAugmentReq(BaseModel):
    existing_runs: list[dict]
    factors: list[str]
    mode: str = "center"      # center | axial | fold | replicate
    n_center: int = 4
    alpha: float | None = None
    n_replicates: int = 1


@app.post("/stats/doe-augment")
def doe_augment(req: DoeAugmentReq):
    return doe_stat.augment(req.existing_runs, req.factors,
                            mode=req.mode, n_center=req.n_center,
                            alpha=req.alpha, n_replicates=req.n_replicates)


class ResponseSurfaceReq(BaseModel):
    rows_storage_key: str
    response: str
    factors: list[str]


@app.post("/stats/response-surface")
def response_surface(req: ResponseSurfaceReq):
    return doe_stat.fit_response_surface(_read_rows(req.rows_storage_key),
                                         response=req.response, factors=req.factors)


class DesirabilityReq(BaseModel):
    rows_storage_key: str
    factors: list[str]
    responses: list[dict]   # [{name, kind, low, high, target?, weight?, importance?}]
    n_starts: int | None = None


@app.post("/stats/desirability")
def desirability(req: DesirabilityReq):
    return doe_stat.multi_response_desirability(
        _read_rows(req.rows_storage_key),
        factors=req.factors, responses=req.responses,
        n_starts=req.n_starts or 24,
    )


class PostHocReq(BaseModel):
    rows_storage_key: str
    test: str
    value_col: str
    group_col: str
    control_group: str | None = None
    direction: str | None = None  # for hsu_mcb: best_is_largest / best_is_smallest
    p_adjust: str | None = None   # for dunn: holm (default) / bonferroni
    alpha: float = 0.05


@app.post("/stats/posthoc")
def posthoc(req: PostHocReq):
    df = _read_rows(req.rows_storage_key)
    if req.test == "tukey_hsd":
        return posthoc_stat.tukey_hsd(df, req.value_col, req.group_col, alpha=req.alpha)
    if req.test == "fisher_lsd":
        return posthoc_stat.fisher_lsd(df, req.value_col, req.group_col, alpha=req.alpha)
    if req.test == "games_howell":
        return posthoc_stat.games_howell(df, req.value_col, req.group_col, alpha=req.alpha)
    if req.test == "dunn":
        return posthoc_stat.dunn(df, req.value_col, req.group_col,
                                 alpha=req.alpha, p_adjust=req.p_adjust or "holm")
    if req.test == "dunnett":
        if not req.control_group:
            raise HTTPException(status_code=400, detail="dunnett requires control_group")
        return posthoc_stat.dunnett(df, req.value_col, req.group_col,
                                    control_group=req.control_group, alpha=req.alpha)
    if req.test == "hsu_mcb":
        direction = getattr(req, "direction", None) or "best_is_largest"
        return posthoc_stat.hsu_mcb(df, req.value_col, req.group_col,
                                    direction=direction, alpha=req.alpha)
    raise HTTPException(status_code=400, detail=f"unknown test: {req.test}")


class ToleranceReq(BaseModel):
    rows_storage_key: str
    method: str
    column: str
    p: float = 0.95
    confidence: float = 0.95
    two_sided: bool = True


@app.post("/stats/tolerance")
def tolerance(req: ToleranceReq):
    df = _read_rows(req.rows_storage_key)
    if req.method == "normal":
        return tolerance_stat.normal(df, req.column, p=req.p, conf=req.confidence,
                                     two_sided=req.two_sided)
    if req.method == "nonparametric":
        return tolerance_stat.nonparametric(df, req.column, p=req.p, conf=req.confidence)
    raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")


class ProbabilityCalcReq(BaseModel):
    distribution: str
    mode: str
    x: float | list[float]
    params: dict


@app.post("/stats/probability")
def probability(req: ProbabilityCalcReq):
    return probability_stat.calculator(req.distribution, req.mode, req.x, req.params)


class ProbabilityPlotReq(BaseModel):
    rows_storage_key: str
    column: str
    distribution: str = "normal"


@app.post("/stats/probability-plot")
def probability_plot(req: ProbabilityPlotReq):
    result = probability_stat.probability_plot(_read_rows(req.rows_storage_key),
                                               req.column, distribution=req.distribution)
    if "chart_png" in result:
        key = _put_bytes("charts", "probplot.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class GraphReq(BaseModel):
    rows_storage_key: str
    chart: str
    column: str | None = None
    columns: list[str] | None = None
    group_col: str | None = None
    x_col: str | None = None
    y_col: str | None = None
    time_col: str | None = None
    factor_cols: list[str] | None = None
    # Interaction-plot only:
    response: str | None = None
    factor_a: str | None = None
    factor_b: str | None = None


@app.post("/stats/graph")
def graph(req: GraphReq):
    df = _read_rows(req.rows_storage_key)
    if req.chart == "boxplot":
        result = graphs_stat.boxplot(df, req.column, req.group_col)
    elif req.chart == "histogram":
        result = graphs_stat.histogram(df, req.column)
    elif req.chart == "scatter":
        result = graphs_stat.scatter(df, req.x_col, req.y_col, req.group_col)
    elif req.chart == "matrix_plot":
        result = graphs_stat.matrix_plot(df, req.columns or [])
    elif req.chart == "time_series":
        result = graphs_stat.time_series_plot(df, req.column, req.time_col)
    elif req.chart == "individual_value_plot":
        result = graphs_stat.individual_value_plot(df, req.column, req.group_col)
    elif req.chart == "run_chart":
        result = graphs_stat.run_chart(df, req.column, req.time_col)
    elif req.chart == "multi_vari":
        result = graphs_stat.multi_vari(df, req.column, req.factor_cols or [])
    elif req.chart == "interaction":
        if not req.response or not req.factor_a or not req.factor_b:
            raise HTTPException(status_code=400, detail="interaction_plot requires response, factor_a, factor_b")
        result = graphs_stat.interaction_plot(df, response=req.response,
                                              factor_a=req.factor_a,
                                              factor_b=req.factor_b)
    else:
        raise HTTPException(status_code=400, detail=f"unknown chart: {req.chart}")
    if "chart_png" in result:
        key = _put_bytes("charts", f"{req.chart}.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class AttributeCapReq(BaseModel):
    rows_storage_key: str
    method: str
    defects_col: str
    n_col: str
    target: float | None = None


@app.post("/stats/attribute-capability")
def attribute_capability(req: AttributeCapReq):
    df = _read_rows(req.rows_storage_key)
    if req.method == "binomial":
        result = attribute_capability_stat.binomial(df, req.defects_col, req.n_col,
                                                    target_p=req.target)
    elif req.method == "poisson":
        result = attribute_capability_stat.poisson(df, req.defects_col, req.n_col,
                                                   target_dpu=req.target)
    else:
        raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")
    if "chart_png" in result:
        key = _put_bytes("charts", "attr_capability.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class AnomReq(BaseModel):
    rows_storage_key: str
    value_col: str
    group_col: str
    alpha: float = 0.05


@app.post("/stats/anom")
def anom(req: AnomReq):
    result = anom_stat.compute(_read_rows(req.rows_storage_key), req.value_col,
                               req.group_col, alpha=req.alpha)
    if "chart_png" in result:
        key = _put_bytes("charts", "anom.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class SixpackReq(BaseModel):
    rows_storage_key: str
    column: str
    lsl: float | None = None
    usl: float | None = None
    target: float | None = None
    subgroup_col: str | None = None


@app.post("/stats/sixpack")
def sixpack(req: SixpackReq):
    result = sixpack_stat.compute(_read_rows(req.rows_storage_key), req.column,
                                  lsl=req.lsl, usl=req.usl,
                                  subgroup_col=req.subgroup_col, target=req.target)
    if "chart_png" in result:
        key = _put_bytes("charts", "sixpack.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class AcceptanceSamplingReq(BaseModel):
    method: str
    aql: float | None = None
    rql: float | None = None
    alpha: float = 0.05
    beta: float = 0.10
    lot_size: int | None = None
    n: int | None = None
    c: int | None = None
    # Variables-plan (Z1.9 / MIL-STD-414) extras
    inspection_level: str | None = None
    sd_known: bool | None = None


@app.post("/stats/acceptance-sampling")
def acceptance_sampling(req: AcceptanceSamplingReq):
    if req.method == "design":
        return acceptance_sampling_stat.design_plan(aql=req.aql, rql=req.rql,
                                                    alpha=req.alpha, beta=req.beta,
                                                    lot_size=req.lot_size)
    if req.method == "oc_curve":
        if req.n is None or req.c is None:
            raise HTTPException(status_code=400, detail="oc_curve requires n and c")
        result = acceptance_sampling_stat.oc_curve(req.n, req.c)
        if "chart_png" in result:
            key = _put_bytes("charts", "oc.png", result.pop("chart_png"), "image/png")
            result["chart_storage_key"] = key
        return result
    if req.method == "variables_z1_9":
        if req.aql is None or req.lot_size is None:
            raise HTTPException(status_code=400, detail="variables_z1_9 requires aql and lot_size")
        return acceptance_sampling_stat.variables_plan_mil_std_414(
            aql=req.aql, lot_size=req.lot_size,
            inspection_level=req.inspection_level or "II",
            sd_known=bool(req.sd_known) if req.sd_known is not None else False)
    raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")


class RandomDataReq(BaseModel):
    distribution: str
    n: int
    params: dict
    seed: int | None = None


@app.post("/stats/random-data")
def random_data(req: RandomDataReq):
    return random_data_stat.generate(req.distribution, req.n, req.params, seed=req.seed)


# ─── New endpoints: Bench-only quality additions ───────────────────────

class AgreementReq(BaseModel):
    rows_storage_key: str
    appraiser_col: str
    part_col: str
    rating_col: str
    standard_col: str | None = None
    trial_col: str | None = None
    ordinal: bool = False


@app.post("/stats/agreement")
def agreement(req: AgreementReq):
    return agreement_stat.compute(_read_rows(req.rows_storage_key),
                                  appraiser_col=req.appraiser_col,
                                  part_col=req.part_col,
                                  rating_col=req.rating_col,
                                  standard_col=req.standard_col,
                                  trial_col=req.trial_col,
                                  ordinal=req.ordinal)


class BootstrapReq(BaseModel):
    rows_storage_key: str
    column: str
    statistic: str = "mean"
    n_boot: int = 5000
    method: str = "bca"
    alpha: float = 0.05
    group_col: str | None = None
    seed: int | None = None


@app.post("/stats/bootstrap")
def bootstrap(req: BootstrapReq):
    return bootstrap_stat.compute(_read_rows(req.rows_storage_key),
                                  column=req.column,
                                  statistic=req.statistic,
                                  n_boot=req.n_boot,
                                  method=req.method,
                                  alpha=req.alpha,
                                  group_col=req.group_col,
                                  seed=req.seed)


class CorrelationReq(BaseModel):
    rows_storage_key: str
    columns: list[str] | None = None
    method: str = "pearson"
    alpha: float = 0.05
    min_r: float = 0.3


@app.post("/stats/correlation")
def correlation(req: CorrelationReq):
    return correlation_stat.compute(_read_rows(req.rows_storage_key),
                                    columns=req.columns,
                                    method=req.method,
                                    alpha=req.alpha,
                                    min_r=req.min_r)


class GageLinearityReq(BaseModel):
    rows_storage_key: str
    part_col: str
    reference_col: str
    measurement_col: str
    process_variation: float | None = None


@app.post("/stats/gage-linearity")
def gage_linearity(req: GageLinearityReq):
    return gage_linearity_stat.compute(_read_rows(req.rows_storage_key),
                                       part_col=req.part_col,
                                       reference_col=req.reference_col,
                                       measurement_col=req.measurement_col,
                                       process_variation=req.process_variation)


# ─── Data wrangling — in-app column transforms ─────────────────────────
#
# The Minitab gap: in-app data manipulation. Users currently re-upload after
# every column change. These endpoints take a rows_storage_key, apply a
# transform, and write back to a NEW storage_key — so the original dataset
# is never mutated and recipes can replay the chain.

class TransformReq(BaseModel):
    rows_storage_key: str
    op: str                            # compute | recode | retype | rename | drop | stack | unstack | impute | filter
    params: dict                       # op-specific


class SurvivalReq(BaseModel):
    rows_storage_key: str
    time_col: str
    event_col: str
    group_col: str | None = None


@app.post("/stats/survival")
def survival(req: SurvivalReq):
    result = survival_stat.kaplan_meier(_read_rows(req.rows_storage_key),
                                        time_col=req.time_col,
                                        event_col=req.event_col,
                                        group_col=req.group_col)
    if "chart_png" in result:
        key = _put_bytes("charts", "survival.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class MixedEffectsReq(BaseModel):
    rows_storage_key: str
    fixed: str
    group: str
    random: str = "1"
    reml: bool = True
    method: str = "lmm"               # lmm | gee | glmm
    family: str | None = None         # gee/glmm: gaussian|binomial|poisson|gamma
    cov_struct: str | None = None     # gee: exchangeable|independence|ar1|unstructured


@app.post("/stats/mixed-effects")
def mixed_effects(req: MixedEffectsReq):
    df = _read_rows(req.rows_storage_key)
    method = (req.method or "lmm").lower()
    if method == "gee":
        return mixed_effects_stat.gee(df, fixed=req.fixed, group=req.group,
                                      family=req.family or "gaussian",
                                      cov_struct=req.cov_struct or "exchangeable")
    if method == "glmm":
        return mixed_effects_stat.glmm(df, fixed=req.fixed, group=req.group,
                                       family=req.family or "binomial")
    return mixed_effects_stat.compute(df, fixed=req.fixed, group=req.group,
                                      random=req.random, reml=req.reml)


class CostParetoReq(BaseModel):
    rows_storage_key: str
    category_col: str
    cost_col: str
    count_col: str | None = None


@app.post("/stats/cost-weighted-pareto")
def cost_weighted_pareto(req: CostParetoReq):
    from stats import pareto as pareto_stat
    result = pareto_stat.cost_weighted(_read_rows(req.rows_storage_key),
                                       category_col=req.category_col,
                                       cost_col=req.cost_col,
                                       count_col=req.count_col)
    if "chart_png" in result:
        key = _put_bytes("charts", "cost_pareto.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class TernaryReq(BaseModel):
    rows_storage_key: str
    components: list[str]
    response: str


@app.post("/stats/ternary-contour")
def ternary_contour(req: TernaryReq):
    result = doe_stat.ternary_contour(_read_rows(req.rows_storage_key),
                                      components=req.components,
                                      response=req.response)
    if "chart_png" in result:
        key = _put_bytes("charts", "ternary.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class BootstrapEffectReq(BaseModel):
    rows_storage_key: str
    column: str
    group_col: str
    kind: str = "cohens_d"
    n_boot: int = 2000
    alpha: float = 0.05
    seed: int | None = None


@app.post("/stats/bootstrap-effect-size")
def bootstrap_effect_size(req: BootstrapEffectReq):
    return bootstrap_stat.effect_size(_read_rows(req.rows_storage_key),
                                      column=req.column,
                                      group_col=req.group_col,
                                      kind=req.kind, n_boot=req.n_boot,
                                      alpha=req.alpha, seed=req.seed)


class PreflightReq(BaseModel):
    rows_storage_key: str
    kind: str
    params: dict


@app.post("/preflight")
def preflight(req: PreflightReq):
    return preflight_stat.check(_read_rows(req.rows_storage_key),
                                kind=req.kind, params=req.params)


class NarrativeReq(BaseModel):
    kind: str
    summary: dict


@app.post("/narrative")
def narrative(req: NarrativeReq):
    """Pure function — no dataset needed. Given a result summary, return the
    decision-grade headline + subhead + verdict."""
    return narrative_stat.for_kind(req.kind, req.summary)


class FollowupsReq(BaseModel):
    kind: str
    summary: dict
    request: dict | None = None


@app.post("/followups")
def followups(req: FollowupsReq):
    """Given a finished result, return suggested next-step analyses."""
    return {"followups": followups_stat.for_kind(req.kind, req.summary,
                                                  request=req.request or {})}


class VariabilityGaugeReq(BaseModel):
    rows_storage_key: str
    measurement_col: str
    part_col: str
    operator_col: str | None = None


@app.post("/stats/variability-gauge")
def variability_gauge(req: VariabilityGaugeReq):
    result = graphs_stat.variability_gauge(_read_rows(req.rows_storage_key),
                                            measurement_col=req.measurement_col,
                                            part_col=req.part_col,
                                            operator_col=req.operator_col)
    if "chart_png" in result:
        key = _put_bytes("charts", "var_gauge.png", result.pop("chart_png"), "image/png")
        result["chart_storage_key"] = key
    return result


class ArlDesignReq(BaseModel):
    chart_kind: str                       # 'cusum' | 'ewma'
    target_arl0: float = 370.4
    shift: float = 1.0
    lam: float | None = None


@app.post("/stats/arl-design")
def arl_design(req: ArlDesignReq):
    return qh_stat.arl_design(chart_kind=req.chart_kind,
                              target_arl0=req.target_arl0,
                              shift=req.shift, lam=req.lam)


class ClementsReq(BaseModel):
    rows_storage_key: str
    column: str
    lsl: float | None = None
    usl: float | None = None
    target: float | None = None


@app.post("/stats/clements-capability")
def clements(req: ClementsReq):
    return qh_stat.clements_capability(_read_rows(req.rows_storage_key),
                                       column=req.column,
                                       lsl=req.lsl, usl=req.usl,
                                       target=req.target)


class DiscreteProbReq(BaseModel):
    distribution: str
    params: dict
    x: float | None = None


@app.post("/stats/discrete-probability")
def discrete_probability(req: DiscreteProbReq):
    return qh_stat.discrete_probability(req.distribution, req.params, x=req.x)


class MixtureEMReq(BaseModel):
    rows_storage_key: str
    column: str
    n_components: int = 2
    max_iter: int = 200
    seed: int | None = 42


@app.post("/stats/mixture-em")
def mixture_em(req: MixtureEMReq):
    return qh_stat.mixture_em(_read_rows(req.rows_storage_key),
                              column=req.column,
                              n_components=req.n_components,
                              max_iter=req.max_iter, seed=req.seed)


class StabilityReq(BaseModel):
    rows_storage_key: str
    time_col: str
    value_col: str
    spec_low: float | None = None
    spec_high: float | None = None
    confidence: float = 0.95


@app.post("/stats/stability")
def stability(req: StabilityReq):
    return qh_stat.stability_regression(_read_rows(req.rows_storage_key),
                                        time_col=req.time_col,
                                        value_col=req.value_col,
                                        spec_low=req.spec_low,
                                        spec_high=req.spec_high,
                                        confidence=req.confidence)


class StressStrengthReq(BaseModel):
    stress_mean: float
    stress_sd: float
    strength_mean: float
    strength_sd: float


@app.post("/stats/stress-strength")
def stress_strength(req: StressStrengthReq):
    return reliability_stat.stress_strength(req.stress_mean, req.stress_sd,
                                            req.strength_mean, req.strength_sd)


class BayesianReq(BaseModel):
    rows_storage_key: str
    method: str                          # 'beta_binomial' | 'normal_normal' | 'best_two_sample' | 'bayes_factor_ttest'
    column: str
    group_col: str | None = None
    # method-specific
    prior_alpha: float | None = None
    prior_beta: float | None = None
    prior_mean: float | None = None
    prior_se: float | None = None
    mu0: float | None = None
    n_draws: int | None = None
    seed: int | None = None
    r: float | None = None
    hdi: float | None = None


@app.post("/stats/bayesian")
def bayesian(req: BayesianReq):
    df = _read_rows(req.rows_storage_key)
    if req.method == "beta_binomial":
        return bayesian_stat.beta_binomial(df, column=req.column,
                                           prior_alpha=req.prior_alpha or 1.0,
                                           prior_beta=req.prior_beta or 1.0,
                                           hdi=req.hdi or 0.95)
    if req.method == "normal_normal":
        return bayesian_stat.normal_normal(df, column=req.column,
                                           prior_mean=req.prior_mean or 0.0,
                                           prior_se=req.prior_se,
                                           hdi=req.hdi or 0.95)
    if req.method == "best_two_sample":
        if not req.group_col:
            raise HTTPException(status_code=400, detail="best_two_sample requires group_col")
        return bayesian_stat.best_two_sample(df, column=req.column,
                                             group_col=req.group_col,
                                             n_draws=req.n_draws or 20_000,
                                             hdi=req.hdi or 0.95,
                                             seed=req.seed)
    if req.method == "bayes_factor_ttest":
        return bayesian_stat.bayes_factor_ttest(df, column=req.column,
                                                group_col=req.group_col,
                                                mu0=req.mu0 or 0.0,
                                                r=req.r or 0.707)
    raise HTTPException(status_code=400, detail=f"unknown method: {req.method}")


class XlsxExportReq(BaseModel):
    kind: str
    params: dict
    summary: dict
    provenance: dict | None = None


@app.post("/export/xlsx")
def export_xlsx(req: XlsxExportReq):
    """Build an .xlsx workbook from an analysis result. Sheets are auto-
    generated based on what's in the summary: a Summary sheet always, plus
    one extra sheet per recognized table (coefficients, ANOVA, per-class,
    rule violations, etc.). Returns the bytes for the Node side to stream."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    accent_fill = PatternFill("solid", fgColor="C5A572")

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Conyso Bench analysis export"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([""])
    ws.append(["Kind", req.kind])
    ws.append(["Parameters", str(req.params)])
    ws.append([""])
    ws.append(["Metric", "Value"])
    for c in ws[ws.max_row]:
        c.font = bold; c.fill = accent_fill
    for k, v in req.summary.items():
        # Flatten confidence-interval dicts (cpk_ci, ppk_ci, …) into a readable
        # "X 95% CI" row instead of dropping them with the other dicts/tables.
        if isinstance(v, dict) and "lo" in v and "hi" in v:
            conf = int(v.get("conf", 0.95) * 100)
            ws.append([f"{k.replace('_ci','')} {conf}% CI", f"[{v['lo']:.4g}, {v['hi']:.4g}]"])
            continue
        if isinstance(v, (list, dict)):
            continue          # table data → own sheet
        ws.append([k, v if v is not None else ""])
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    def _write_table(title: str, rows: list[dict]):
        if not rows:
            return
        wsx = wb.create_sheet(title[:30])     # Excel sheet name limit
        cols = list(rows[0].keys())
        wsx.append(cols)
        for c in wsx[1]:
            c.font = bold; c.fill = accent_fill
        for r in rows:
            wsx.append([r.get(c) if not isinstance(r.get(c), (list, dict)) else str(r.get(c))
                        for c in cols])
        for i, _ in enumerate(cols, 1):
            wsx.column_dimensions[get_column_letter(i)].width = 18

    # Recognise common tabular fields and give each its own sheet.
    s = req.summary
    if isinstance(s.get("coefficients"), list):       _write_table("Coefficients", s["coefficients"])
    if isinstance(s.get("table"), list):              _write_table("ANOVA table", s["table"])
    if isinstance(s.get("per_class"), list):          _write_table("Per-class metrics", s["per_class"])
    if isinstance(s.get("rule_violations"), list):    _write_table("Rule violations", s["rule_violations"])
    if isinstance(s.get("compact_letter_display"), list):
        _write_table("Tukey letters", s["compact_letter_display"])
    if isinstance(s.get("feature_importance"), list): _write_table("Feature importance", s["feature_importance"])
    if isinstance(s.get("comparisons"), list):        _write_table("Pairwise", s["comparisons"])
    if isinstance(s.get("per_part"), list):           _write_table("Per-part bias", s["per_part"])
    if isinstance(s.get("significant"), list):        _write_table("Significant pairs", s["significant"])
    if isinstance(s.get("by_frequency"), list):       _write_table("By frequency", s["by_frequency"])
    if isinstance(s.get("by_cost"), list):            _write_table("By cost", s["by_cost"])
    if isinstance(s.get("segments"), list):           _write_table("Segments", s["segments"])

    # Provenance footer.
    if req.provenance:
        wsx = wb.create_sheet("Provenance")
        wsx.append(["Key", "Hash"])
        for c in wsx[1]: c.font = bold; c.fill = accent_fill
        for k, v in req.provenance.items():
            wsx.append([k, v])
        wsx.column_dimensions["A"].width = 22
        wsx.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    bytes_ = buf.getvalue()
    # Persist as a one-shot file and return its key — Node will stream it.
    key = _put_bytes("files", f"export-{uuid.uuid4()}.xlsx", bytes_,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return {"storage_key": key, "n_bytes": len(bytes_)}


@app.post("/wrangle/transform")
def wrangle_transform(req: TransformReq):
    df = _read_rows(req.rows_storage_key)
    out_df, meta = transform_mod.apply(df, op=req.op, params=req.params)
    # Reuse /materialize-rows logic — write parquet, return key + schema.
    sub = ROWS_DIR / str(uuid.uuid4())
    sub.mkdir(parents=True, exist_ok=True)
    out_path = sub / "rows.parquet"
    out_df.to_parquet(out_path, index=False)
    schema = []
    for col in out_df.columns:
        t = out_df[col].dtype.name
        schema.append({
            "name": col,
            "type": "number" if "int" in t or "float" in t else ("date" if "datetime" in t else "string"),
            "n_unique": int(out_df[col].nunique()),
            "n_null": int(out_df[col].isna().sum()),
        })
    return {"summary": meta,
            "materialized": {
                "rows_storage_key": str(out_path.relative_to(DATA_DIR)),
                "n_rows": int(len(out_df)),
                "schema": schema,
            }}


# ─── Static file delivery ──────────────────────────────────────────────
#
# Charts and uploaded artifacts live on disk under DATA_DIR. The Node
# server proxies /artifact/<storage_key> requests here.

@app.get("/file/{path:path}")
def get_file(path: str):
    from fastapi.responses import FileResponse
    p = (DATA_DIR / path).resolve()
    if not str(p).startswith(str(DATA_DIR)):
        raise HTTPException(status_code=400, detail="invalid path")
    if not p.exists():
        raise HTTPException(status_code=404, detail="not found")
    return FileResponse(p)


# ─── Upload (handled here directly — no separate Node piece) ───────────

from fastapi import UploadFile, File


@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    """Save the uploaded file under data/files/<uuid>/<name> and return
    the storage_key. Node side will then call /parse/* with that key."""
    sub = FILES_DIR / str(uuid.uuid4())
    sub.mkdir(parents=True, exist_ok=True)
    out_path = sub / (file.filename or "upload.bin")
    contents = await file.read()
    out_path.write_bytes(contents)
    return {
        "storage_key": str(out_path.relative_to(DATA_DIR)),
        "filename": file.filename,
        "size_bytes": len(contents),
        "content_type": file.content_type,
    }


@app.post("/materialize-rows")
def materialize_rows(payload: dict):
    """Take parsed rows (list[dict]) → save as a parquet file → return
    rows_storage_key + schema. Lightweight standalone version of the
    full-product wrangler.materialize.
    """
    import pandas as pd
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(status_code=400, detail="rows must be a non-empty list")
    df = pd.DataFrame(rows)
    sub = ROWS_DIR / str(uuid.uuid4())
    sub.mkdir(parents=True, exist_ok=True)
    out_path = sub / "rows.parquet"
    df.to_parquet(out_path, index=False)
    schema = []
    for col in df.columns:
        t = df[col].dtype.name
        schema.append({
            "name": col,
            "type": "number" if "int" in t or "float" in t else ("date" if "datetime" in t else "string"),
            "n_unique": int(df[col].nunique()),
            "n_null": int(df[col].isna().sum()),
        })
    return {
        "rows_storage_key": str(out_path.relative_to(DATA_DIR)),
        "n_rows": int(len(df)),
        "schema": schema,
    }


@app.get("/healthz")
def healthz():
    return {"ok": True}
